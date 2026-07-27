#!/usr/bin/env python3
"""Sincroniza el padrón oficial argentino con public.schools en Supabase."""

from __future__ import annotations

import html
import os
import re
import sys
import tempfile
import time
import unicodedata
import uuid
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook


PADRON_PAGE = (
    "https://www.argentina.gob.ar/educacion/"
    "evaluacion-e-informacion-educativa/"
    "padron-oficial-de-establecimientos-educativos"
)
SOURCE_NAME = "official_padron"
REQUEST_TIMEOUT = 90
UPSERT_BATCH_SIZE = 350
MAX_RETRIES = 4

# Índices de columnas que corresponden a nivel inicial, primario o secundario.
LEVEL_COLUMNS = {
    "Inicial": (16, 17, 24, 25, 34),
    "Primaria": (18, 26, 29, 35),
    "Secundaria": (19, 20, 27, 30, 36),
}


class LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[str] = []

    def handle_starttag(
        self,
        tag: str,
        attrs: list[tuple[str, str | None]],
    ) -> None:
        if tag.lower() != "a":
            return

        href = dict(attrs).get("href")
        if href:
            self.links.append(html.unescape(href))


def require_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise RuntimeError(f"Falta configurar el secreto {name}.")
    return value


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def compact(value: Any, max_length: int | None = None) -> str | None:
    if value is None:
        return None

    result = re.sub(r"\s+", " ", str(value)).strip()
    if not result:
        return None

    if max_length:
        result = result[:max_length]
    return result


def pretty_text(value: Any) -> str | None:
    text = compact(value)
    if not text:
        return None

    if text == text.upper():
        text = text.title()

    replacements = {
        " De ": " de ",
        " Del ": " del ",
        " La ": " la ",
        " Las ": " las ",
        " Los ": " los ",
        " Y ": " y ",
        " E ": " e ",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)

    return text


def normalized(value: Any) -> str:
    text = unicodedata.normalize("NFD", compact(value) or "")
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def is_marked(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() not in {"", "0", "no", "false", "none"}


def discover_latest_xlsx(session: requests.Session) -> str:
    response = session.get(PADRON_PAGE, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()

    parser = LinkParser()
    parser.feed(response.text)
    links = [urljoin(PADRON_PAGE, link) for link in parser.links]

    current_pages = [
        link
        for link in links
        if "/files/" in link.lower() and "padron" in link.lower()
    ]

    for page_url in current_pages:
        detail = session.get(page_url, timeout=REQUEST_TIMEOUT)
        detail.raise_for_status()

        detail_parser = LinkParser()
        detail_parser.feed(detail.text)

        for link in detail_parser.links:
            resolved = urljoin(page_url, link)
            if resolved.lower().split("?")[0].endswith(".xlsx"):
                return resolved

    direct_files = [
        link
        for link in links
        if link.lower().split("?")[0].endswith(".xlsx")
        and "padron" in link.lower()
    ]
    if direct_files:
        return direct_files[0]

    raise RuntimeError("No se encontró el archivo XLSX del padrón oficial.")


def download_file(
    session: requests.Session,
    url: str,
    destination: Path,
) -> None:
    with session.get(url, stream=True, timeout=REQUEST_TIMEOUT) as response:
        response.raise_for_status()
        with destination.open("wb") as output:
            for chunk in response.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    output.write(chunk)


def detect_header_row(worksheet: Any) -> int:
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=40, values_only=True),
        start=1,
    ):
        values = {compact(value) for value in row if compact(value)}
        if {"Jurisdicción", "Cueanexo", "Nombre"}.issubset(values):
            return row_number

    raise RuntimeError("No se encontró la cabecera esperada en el padrón.")


def official_code(cue_root: str) -> str:
    # Código estable, corto y único. Ejemplo: CUE 0200001 -> AR020-0001.
    return f"AR{cue_root[:3]}-{cue_root[3:]}"


def levels_for_row(row: tuple[Any, ...]) -> set[str]:
    result: set[str] = set()
    for level, indexes in LEVEL_COLUMNS.items():
        if any(index < len(row) and is_marked(row[index]) for index in indexes):
            result.add(level)
    return result


def build_school(
    row: tuple[Any, ...],
    cue: str,
    cue_root: str,
    levels: set[str],
    run_id: str,
    source_url: str,
    synced_at: str,
) -> dict[str, Any]:
    province = pretty_text(row[0]) or "Argentina"
    city = pretty_text(row[5])
    department = pretty_text(row[3])

    return {
        "code": official_code(cue_root),
        "cue": cue,
        "cue_root": cue_root,
        "name": pretty_text(row[8]) or f"Establecimiento {cue}",
        "province": province,
        "city": city,
        "zone_code": city or department or province,
        "department": department,
        "locality_code": compact(row[6], 30),
        "address": pretty_text(row[9]),
        "postal_code": compact(row[10], 20),
        "phone": compact(row[11], 160),
        "email": compact(row[12], 320),
        "sector": pretty_text(row[1]),
        "geographic_area": pretty_text(row[2]),
        "education_levels": sorted(levels),
        "is_active": True,
        "source": SOURCE_NAME,
        "source_url": source_url,
        "source_updated_at": synced_at,
        "sync_run_id": run_id,
    }


def read_schools(
    xlsx_path: Path,
    run_id: str,
    source_url: str,
    synced_at: str,
) -> tuple[list[dict[str, Any]], int]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = detect_header_row(worksheet)

    schools: dict[str, dict[str, Any]] = {}
    source_rows = 0

    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        source_rows += 1

        cue = re.sub(r"\D", "", compact(row[7]) or "")
        if 1 <= len(cue) <= 9:
            cue = cue.zfill(9)
        name = compact(row[8])
        levels = levels_for_row(row)

        if len(cue) != 9 or not name or not levels:
            continue

        cue_root = cue[:7]
        current = schools.get(cue_root)

        if not current:
            schools[cue_root] = build_school(
                row,
                cue,
                cue_root,
                levels,
                run_id,
                source_url,
                synced_at,
            )
            continue

        combined_levels = set(current["education_levels"])
        combined_levels.update(levels)
        current["education_levels"] = sorted(combined_levels)

        # La sede terminada en 00 es la referencia principal de la institución.
        if cue.endswith("00") and not str(current["cue"]).endswith("00"):
            replacement = build_school(
                row,
                cue,
                cue_root,
                combined_levels,
                run_id,
                source_url,
                synced_at,
            )
            schools[cue_root] = replacement

    workbook.close()
    return list(schools.values()), source_rows


class SupabaseRest:
    def __init__(self, url: str, service_key: str) -> None:
        self.base_url = f"{url.rstrip('/')}/rest/v1"
        self.session = requests.Session()
        self.session.headers.update(
            {
                "apikey": service_key,
                "Authorization": f"Bearer {service_key}",
                "Content-Type": "application/json",
            }
        )

    def request(
        self,
        method: str,
        path: str,
        *,
        params: dict[str, str] | None = None,
        payload: Any = None,
        prefer: str | None = None,
    ) -> requests.Response:
        headers = {"Prefer": prefer} if prefer else None
        last_error: Exception | None = None

        for attempt in range(1, MAX_RETRIES + 1):
            try:
                response = self.session.request(
                    method,
                    f"{self.base_url}/{path.lstrip('/')}",
                    params=params,
                    json=payload,
                    headers=headers,
                    timeout=REQUEST_TIMEOUT,
                )

                if response.status_code < 500:
                    response.raise_for_status()
                    return response

                response.raise_for_status()
            except requests.RequestException as error:
                last_error = error
                if attempt == MAX_RETRIES:
                    break
                time.sleep(2 ** (attempt - 1))

        raise RuntimeError(f"Falló la comunicación con Supabase: {last_error}")

    def create_run(self, run_id: str) -> None:
        self.request(
            "POST",
            "school_sync_runs",
            payload={
                "id": run_id,
                "status": "running",
                "source_url": PADRON_PAGE,
            },
            prefer="return=minimal",
        )

    def update_run(self, run_id: str, values: dict[str, Any]) -> None:
        self.request(
            "PATCH",
            "school_sync_runs",
            params={"id": f"eq.{run_id}"},
            payload=values,
            prefer="return=minimal",
        )

    def upsert_schools(self, schools: list[dict[str, Any]]) -> None:
        total = len(schools)

        for start in range(0, total, UPSERT_BATCH_SIZE):
            batch = schools[start : start + UPSERT_BATCH_SIZE]
            self.request(
                "POST",
                "schools",
                params={"on_conflict": "cue_root"},
                payload=batch,
                prefer="resolution=merge-duplicates,return=minimal",
            )

            imported = min(start + len(batch), total)
            print(f"Importados {imported}/{total} colegios.")

    def finalize_sync(self, run_id: str) -> int:
        response = self.request(
            "POST",
            "rpc/finalize_school_sync",
            payload={"p_run_id": run_id},
        )
        return int(response.json() or 0)
def main() -> int:
    supabase_url = require_env("SUPABASE_URL")
    service_key = require_env("SUPABASE_SERVICE_ROLE_KEY")
    run_id = str(uuid.uuid4())
    synced_at = utc_now()

    rest = SupabaseRest(supabase_url, service_key)
    source_session = requests.Session()
    source_session.headers.update(
        {
            "User-Agent": (
                "ColegioLibreSchoolSync/1.0 "
                "(contacto: ayudacolegiolibre@gmail.com)"
            )
        }
    )

    rest.create_run(run_id)

    try:
        source_url = discover_latest_xlsx(source_session)
        rest.update_run(run_id, {"source_url": source_url})
        print(f"Padrón encontrado: {source_url}")

        with tempfile.TemporaryDirectory(prefix="colegiolibre-schools-") as temp:
            xlsx_path = Path(temp) / "padron.xlsx"
            download_file(source_session, source_url, xlsx_path)
            schools, source_rows = read_schools(
                xlsx_path,
                run_id,
                source_url,
                synced_at,
            )

        if len(schools) < 1000:
            raise RuntimeError(
                "El padrón produjo muy pocos colegios; se canceló para evitar "
                "desactivar datos válidos."
            )

        print(
            f"Se procesaron {source_rows} filas y "
            f"{len(schools)} instituciones escolares."
        )

        rest.upsert_schools(schools)
        deactivated = rest.finalize_sync(run_id)

        rest.update_run(
            run_id,
            {
                "status": "completed",
                "finished_at": utc_now(),
                "source_rows": source_rows,
                "imported_schools": len(schools),
                "deactivated_schools": deactivated,
                "error_message": None,
            },
        )

        print(
            "Sincronización terminada: "
            f"{len(schools)} colegios activos y "
            f"{deactivated} desactivados."
        )
        return 0

    except Exception as error:
        message = compact(error, 1000) or "Error desconocido."
        print(f"ERROR: {message}", file=sys.stderr)

        try:
            rest.update_run(
                run_id,
                {
                    "status": "failed",
                    "finished_at": utc_now(),
                    "error_message": message,
                },
            )
        except Exception as update_error:
            print(
                f"No se pudo registrar el error: {update_error}",
                file=sys.stderr,
            )

        return 1


if __name__ == "__main__":
    raise SystemExit(main())
